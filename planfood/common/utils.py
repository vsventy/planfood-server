import datetime
import logging
from copy import copy

from django.http import HttpResponse

import holidays
from openpyxl.styles import Border, Side
from openpyxl.styles.alignment import Alignment

logger = logging.getLogger(__name__)

def trigger_error(request):
    division_by_zero = 1 / 0

def set_cells_border(worksheet, row, start_column, end_column, border=None):
    if border is None:
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin'),
        )
    for column in range(start_column, end_column):
        cell = worksheet.cell(row=row, column=column)
        cell.border = border


def get_cell_by_value(worksheet, value, min_row=1, max_col=20, max_row=24):
    for row in worksheet.iter_rows(min_row=min_row, max_col=max_col, max_row=max_row):
        for cell in row:
            if cell.value == value or cell.value == '{{{}}}'.format(value):
                return cell
    return None


def clone_cell_style(source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.number_format = copy(source.number_format)
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def set_text_style(cell, name='Arial', size=10.0, bold=False, italic=False):
    cell.font = cell.font.copy(name=name, bold=bold, italic=italic, sz=size)


def set_alignment(cell, indent=0):
    cell.alignment = cell.alignment.copy(indent=indent)


def render_worksheet(worksheet, d, min_row=1, max_column=20, max_row=200):
    for row in worksheet.iter_rows(
        min_row=min_row, max_col=max_column, max_row=max_row
    ):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.format(**d)


def make_spreadsheat_reponse(workbook, filename):
    logger.debug('{}'.format(filename))
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'
    ] = "attachment; filename*=UTF-8''{filename}.xlsx".format(filename=filename)
    workbook.save(response)

    return response


def previous_business_day(date=None):
    if date is None:
        date = datetime.date.today()
    ONE_DAY = datetime.timedelta(days=1)
    previous_day = date - ONE_DAY
    while previous_day.weekday() in holidays.WEEKEND or previous_day in holidays.UA():
        previous_day -= ONE_DAY
    return previous_day


downcase_first_character = lambda s: s[:1].lower() + s[1:] if s else ''
