import json
import os.path
import logging
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from typing import Any, Dict, Type

from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.utils.encoding import escape_uri_path
from django.utils.translation import ugettext_lazy as _
from django.views import View

from openpyxl import load_workbook
from openpyxl.styles import Border, numbers, Side
from openpyxl.styles.alignment import Alignment
from rest_framework.views import APIView

from planfood.common.models import DefaultSettings, Group
from planfood.products.models import Product
from planfood.common.utils import (
    get_cell_by_value,
    clone_cell_style,
    make_spreadsheat_reponse,
    previous_business_day,
    render_worksheet,
    set_cells_border,
)
from .forms import NormsAnalysisForm
from .models import DishItem, MenuDay
from .tasks import create_norms_analysis_report


BASE = os.path.dirname(os.path.abspath(__file__))
MENU_WORKSHEET_NAME = 'Меню'
DEMAND_WORKSHEET_NAME = 'Вимога'

logger = logging.getLogger(__name__)


class ListMenuDays(APIView):
    pass


def init_menu_dict():
    return dict.fromkeys(
        [
            'director_initials',
            'group_name',
            'first_demand_row',
            'first_menu_row',
            'first_norm_header',
            'first_output_header',
            'first_products_header',
            'first_total_age_category',
            'first_total_number',
            'first_issued_header',
            'menu_day',
            'signature_day',
            'nurse_post',
            'nurse_initials',
            'senior_nurse_post',
            'senior_nurse_initials',
            'storekeeper_initials',
        ],
        '',
    )


def create_menu_report_xlsx(request, menuday_id=None):
    report_path = 'spreadsheets/menu_report.xlsx'
    abs_report_path = os.path.abspath(os.path.join(BASE, '../../', report_path))

    if menuday_id is not None:
        menu_day = MenuDay.objects.get(pk=menuday_id)
    elif 'date' in request.GET:
        day = datetime.strptime(request.GET.get('date'), '%d.%m.%Y')
        menu_day = MenuDay.objects.get(date=day)
    else:
        messages.error(request, _('No menu day selected.'))
        return HttpResponseRedirect('/')

    try:
        settings = DefaultSettings.objects.get()
    except DefaultSettings.DoesNotExist:
        messages.error(
            request,
            _('No default settings. Please set them in the admin panel and try again.'),
        )
        return HttpResponseRedirect('/')

    group_name = menu_day.group.name

    workbook = load_workbook(abs_report_path)

    menu_template_cells = get_menu_template_cells(workbook)
    demand_template_cells = get_demand_template_cells(workbook)

    number_of_products = fill_workbook(
        workbook, settings, menu_day, menu_template_cells, demand_template_cells
    )
    format_menu_worksheet(workbook, menu_day, menu_template_cells)
    format_demand_worksheet(workbook, number_of_products, demand_template_cells)

    group_name = group_name.replace(', ', '_').replace(' ', '_')
    filename = '{date}_{group}'.format(
        date=menu_day.date.strftime('%Y-%m-%d'), group=escape_uri_path(group_name)
    )
    return make_spreadsheat_reponse(workbook, filename)


def get_menu_template_cells(workbook):
    menu_worksheet = workbook[MENU_WORKSHEET_NAME]

    return {
        'total_age_category': get_cell_by_value(
            menu_worksheet, 'first_total_age_category'
        ),
        'output_header': get_cell_by_value(menu_worksheet, 'first_output_header'),
        'norm_header': get_cell_by_value(menu_worksheet, 'first_norm_header'),
        'products_header': get_cell_by_value(menu_worksheet, 'first_products_header'),
        'menu_row': get_cell_by_value(menu_worksheet, 'first_menu_row'),
    }


def get_demand_template_cells(workbook):
    demand_worksheet = workbook[DEMAND_WORKSHEET_NAME]

    return {
        'demand_row': get_cell_by_value(demand_worksheet, 'first_demand_row'),
        'group_name': get_cell_by_value(demand_worksheet, 'group_name'),
        'products_header': get_cell_by_value(demand_worksheet, 'first_products_header'),
        'issued_header': get_cell_by_value(demand_worksheet, 'first_issued_header'),
    }


def fill_workbook(workbook, settings, menu_day, menu_cells, demand_cells):
    menu_worksheet = workbook[MENU_WORKSHEET_NAME]
    demand_worksheet = workbook[DEMAND_WORKSHEET_NAME]

    d = init_menu_dict()
    d['menu_day'] = menu_day.date.strftime('%d.%m.%Y')
    d['signature_day'] = previous_business_day(menu_day.date).strftime('%d.%m.%Y')
    d['group_name'] = menu_day.group.name
    d['director_initials'] = settings.director_initials
    d['nurse_post'] = settings.nurse_post
    d['nurse_initials'] = settings.nurse_initials
    d['senior_nurse_post'] = settings.senior_nurse_post
    d['senior_nurse_initials'] = settings.senior_nurse_initials
    d['storekeeper_initials'] = settings.storekeeper_initials

    # numbers of persons by age categories
    total_numbers = menu_day.numbers_of_persons.order_by(
        'age_category__sort'
    ).values_list('age_category__name', 'value')
    total_cell = menu_cells['total_age_category']
    index = 0
    for item in total_numbers:
        menu_worksheet.cell(
            row=total_cell.row + index, column=total_cell.column
        ).value = item[0]
        number_cell = menu_worksheet.cell(
            row=total_cell.row + index, column=total_cell.column + 3
        )
        number_cell.value = item[1]
        number_cell.number_format = numbers.FORMAT_NUMBER
        index += 1

    # init headers of the main tables
    menu_output_header = menu_cells['output_header']
    menu_norm_header = menu_cells['norm_header']
    menu_products_header = menu_cells['products_header']
    demand_products_header = demand_cells['products_header']
    demand_issued_header = demand_cells['issued_header']
    group = menu_day.group
    index = 0
    for age_category in group.age_categories.iterator():
        menu_worksheet.cell(
            row=menu_output_header.row, column=menu_output_header.column + index
        ).value = age_category.name
        menu_worksheet.cell(
            row=menu_norm_header.row, column=menu_norm_header.column + index
        ).value = age_category.name
        menu_worksheet.cell(
            row=menu_products_header.row, column=menu_products_header.column + index
        ).value = age_category.name
        demand_worksheet.cell(
            row=demand_products_header.row, column=demand_products_header.column + index
        ).value = age_category.name
        demand_worksheet.cell(
            row=demand_issued_header.row, column=demand_issued_header.column + index
        ).value = age_category.name
        index += 1

    # fill menu worksheet
    menu_row = menu_cells['menu_row']
    dish_items = menu_day.dish_items.order_by('period')
    products_dict = {}  # type: Dict[Any, Any]
    number_of_products = 0
    index = 0
    for idx, dish_item in enumerate(dish_items):
        # period
        active_cell = menu_worksheet.cell(
            row=menu_row.row + index, column=menu_row.column
        )
        active_cell.value = str(DishItem.PERIOD[dish_item.period])

        # dishes
        n = 0
        for dish in dish_item.dishes.iterator():
            index += 1
            n += 1
            menu_worksheet.insert_rows(menu_row.row + index)
            dish_cell = menu_worksheet.cell(
                row=menu_row.row + index,
                column=menu_row.column,
                value='{}. {}'.format(n, dish.name),
            )

            # dish outputs
            column_index = 1
            for output in (
                dish.outputs.filter(group=group)
                .order_by('age_category__sort')
                .iterator()
            ):
                menu_worksheet.cell(
                    row=dish_cell.row,
                    column=dish_cell.column + column_index,
                    value=str(output.value),
                )
                column_index += 1

            # dish products
            for product in dish.products.iterator():
                if not dish.is_atomic:
                    index += 1
                    menu_worksheet.insert_rows(menu_row.row + index)
                    product_cell = menu_worksheet.cell(
                        row=menu_row.row + index,
                        column=menu_row.column,
                        value=product.name,
                    )
                else:
                    product_cell = dish_cell

                if product.item_number not in products_dict:
                    products_dict[product.item_number] = dict.fromkeys(
                        ['name', 'unit', 'norms']
                    )
                    products_dict[product.item_number]['name'] = product.name
                    products_dict[product.item_number]['unit'] = str(
                        Product.UNIT[product.unit]
                    )

                # product norms
                column_index = 0
                product_norms = dish.norms.filter(
                    group=group, product=product
                ).order_by('age_category__sort')
                total_number_cell = menu_worksheet.cell(
                    row=total_cell.row, column=total_cell.column + 3
                )
                calculated_products_norms = []
                for norm in product_norms.iterator():
                    product_norm_cell = menu_worksheet.cell(
                        row=product_cell.row,
                        column=menu_norm_header.column + column_index,
                        value=norm.value,
                    )
                    menu_worksheet.cell(
                        row=product_cell.row,
                        column=menu_products_header.column + column_index,
                        value='={}{}*{}{}'.format(
                            product_norm_cell.column_letter,
                            product_norm_cell.row,
                            total_number_cell.column_letter,
                            total_number_cell.row + column_index,
                        ),
                    )

                    total_value = menu_worksheet.cell(
                        row=total_cell.row + column_index, column=total_cell.column + 3
                    ).value

                    if total_value is not None:
                        calculated_products_norms.append(
                            product_norm_cell.value * total_value
                        )
                    column_index += 1

                products_norms = products_dict[product.item_number]['norms']
                if products_norms is not None:
                    products_dict[product.item_number]['norms'] = [
                        x + y for x, y in zip(products_norms, calculated_products_norms)
                    ]
                else:
                    products_dict[product.item_number][
                        'norms'
                    ] = calculated_products_norms

                sum_of_products_cell = menu_worksheet.cell(
                    row=product_cell.row, column=menu_products_header.column + 5
                )
                prev_products_cell = menu_worksheet.cell(
                    row=product_cell.row, column=sum_of_products_cell.column - 1
                )
                sum_of_products_cell.value = '=SUM({}{}:{}{})'.format(
                    menu_products_header.column_letter,
                    product_cell.row,
                    prev_products_cell.column_letter,
                    product_cell.row,
                )

        if idx < dish_items.count() - 1:
            index += 1
            menu_worksheet.insert_rows(menu_row.row + index)

    # fill demand worksheet
    sorted_products_dict = dict(sorted(products_dict.items()))
    demand_row = demand_cells['demand_row']
    n = 0
    for item_number, product in sorted_products_dict.items():
        demand_worksheet.cell(
            row=demand_row.row + n, column=demand_row.column, value=str(n + 1)
        )
        demand_worksheet.cell(
            row=demand_row.row + n, column=demand_row.column + 1, value=item_number
        )
        demand_worksheet.cell(
            row=demand_row.row + n, column=demand_row.column + 2, value=product['name']
        )
        demand_worksheet.cell(
            row=demand_row.row + n, column=demand_row.column + 3, value=product['unit']
        )
        column_index = 4
        for norm in product['norms']:
            demand_worksheet.cell(
                row=demand_row.row + n,
                column=demand_row.column + column_index,
                value=norm,
            )
            demand_worksheet.cell(
                row=demand_row.row + n,
                column=demand_row.column + 7 + column_index,
                value=norm,
            )
            column_index += 1
        demand_worksheet.cell(
            row=demand_row.row + n,
            column=demand_row.column + 9,
            value='=SUM(E{}:I{})'.format(  # sum(product['norms'])
                demand_row.row + n, demand_row.row + n
            ),
        )
        demand_worksheet.cell(
            row=demand_row.row + n,
            column=demand_row.column + 16,
            value='=SUM(L{}:P{})'.format(demand_row.row + n, demand_row.row + n),
        )
        n += 1
        number_of_products = len(products_dict)
        if n < number_of_products:
            demand_worksheet.insert_rows(demand_row.row + n)

    render_worksheet(menu_worksheet, d)
    render_worksheet(demand_worksheet, d)

    return number_of_products


def format_menu_worksheet(workbook, menu_day, menu_cells):
    menu_worksheet = workbook[MENU_WORKSHEET_NAME]

    menu_row = menu_cells['menu_row']
    norm_cell = menu_cells['norm_header']
    products_cell = menu_cells['products_header']
    total_cell = menu_cells['total_age_category']
    group = menu_day.group
    dish_items = menu_day.dish_items.order_by('period')
    index = 0
    d_index = 0
    for idx, dish_item in enumerate(dish_items):
        # period
        active_menu_cell = menu_worksheet.cell(
            row=menu_row.row + index, column=menu_row.column
        )
        active_menu_cell.font = active_menu_cell.font.copy(sz=12.0)
        # period borders
        border = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        set_cells_border(
            menu_worksheet,
            active_menu_cell.row,
            2,
            total_cell.column + 6,
            border=border,
        )
        border = Border(
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin'),
        )
        set_cells_border(
            menu_worksheet,
            active_menu_cell.row,
            total_cell.column + 5,
            total_cell.column + 6,
            border=border,
        )

        # dishes
        n = 0
        for dish in dish_item.dishes.iterator():
            index += 1
            n += 1
            dish_cell = menu_worksheet.cell(
                row=menu_row.row + index, column=menu_row.column
            )
            clone_cell_style(active_menu_cell, dish_cell)
            dish_cell.font = dish_cell.font.copy(italic=True, sz=10.0)
            dish_cell.alignment = Alignment(indent=0)

            # dish outputs
            column_index = 1
            for output in dish.outputs.filter(group=group).iterator():
                new_cell = menu_worksheet.cell(
                    row=dish_cell.row, column=dish_cell.column + column_index
                )
                clone_cell_style(active_menu_cell, new_cell)
                new_cell.font = new_cell.font.copy(bold=False, sz=11.0)
                new_cell.alignment = Alignment(indent=0, horizontal='right')
                column_index += 1

            # dish products
            for product in dish.products.iterator():
                if not dish.is_atomic:
                    index += 1
                    product_cell = menu_worksheet.cell(
                        row=menu_row.row + index, column=menu_row.column
                    )
                    clone_cell_style(active_menu_cell, product_cell)
                    product_cell.font = product_cell.font.copy(bold=False, sz=10.0)

                else:
                    product_cell = dish_cell

                # product norms
                column_index = 0
                product_norms = dish.norms.filter(
                    group=group, product=product
                ).order_by('age_category__sort')
                for norm in product_norms.iterator():
                    product_norm_cell = menu_worksheet.cell(
                        row=product_cell.row, column=norm_cell.column + column_index
                    )
                    number_of_products_cell = menu_worksheet.cell(
                        row=product_cell.row, column=products_cell.column + column_index
                    )

                    clone_cell_style(active_menu_cell, product_norm_cell)
                    clone_cell_style(active_menu_cell, number_of_products_cell)
                    product_norm_cell.font = product_norm_cell.font.copy(
                        bold=False, sz=11.0
                    )
                    number_of_products_cell.font = number_of_products_cell.font.copy(
                        bold=False, sz=11.0
                    )
                    product_norm_cell.alignment = Alignment(indent=0)
                    number_of_products_cell.alignment = Alignment(indent=0)
                    column_index += 1

                d_index += 1

                sum_of_products_cell = menu_worksheet.cell(
                    row=product_cell.row, column=products_cell.column + 5
                )
                clone_cell_style(active_menu_cell, sum_of_products_cell)
                sum_of_products_cell.font = new_cell.font.copy(bold=False, sz=11.0)
                sum_of_products_cell.alignment = Alignment(indent=0)

                set_cells_border(
                    menu_worksheet, product_cell.row, 1, total_cell.column + 6
                )

            set_cells_border(menu_worksheet, dish_cell.row, 1, total_cell.column + 6)

        if idx < dish_items.count() - 1:
            index += 1
            new_cell = menu_worksheet.cell(
                row=menu_row.row + index, column=menu_row.column
            )
            clone_cell_style(active_menu_cell, new_cell)


def format_demand_worksheet(workbook, number_of_products, demand_cells):
    demand_worksheet = workbook[DEMAND_WORKSHEET_NAME]

    demand_row = demand_cells['demand_row']
    group_cell = demand_cells['group_name']

    for i in range(number_of_products):
        for j in range(group_cell.column):
            active_demand_cell = demand_worksheet.cell(
                row=demand_row.row + i, column=demand_row.column + j
            )
            clone_cell_style(demand_row, active_demand_cell)
            if (j == 0) or (j == 3):
                active_demand_cell.alignment = Alignment(horizontal='center')
            elif j > 3:
                active_demand_cell.alignment = Alignment(horizontal='right')
            else:
                active_demand_cell.alignment = Alignment(horizontal='left')


class NormsAnalysisView(View):
    def get(self, request):
        form = NormsAnalysisForm(initial={'month': date.today()})
        context = {}
        context['form'] = form
        return render(request, 'menu/home.html', context)

    def post(self, request):
        form = NormsAnalysisForm(request.POST)  # type: Type[NormsAnalysisForm]
        context = {}

        if form.is_valid():
            start_date = form.cleaned_data.get('month')
            end_date = start_date + relativedelta(months=1) - timedelta(days=1)
            group = form.cleaned_data.get('group')
            task = create_norms_analysis_report.delay(
                start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), group.pk
            )
            context['task_id'] = task.id
            context['task_status'] = task.status

            return render(request, 'menu/home.html', context)

        context['form'] = form

        return render(request, 'menu/home.html', context)
