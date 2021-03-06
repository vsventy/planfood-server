import json
import logging
import os
from copy import copy
from datetime import datetime
from typing import Any, Dict

from django.conf import settings
from django.db.models import Q
from django.utils import translation
from django.utils.formats import date_format

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles.alignment import Alignment
from openpyxl.worksheet.properties import PageSetupProperties
from celery.task import task

from planfood.common.models import Group
from planfood.common.utils import (
    clone_cell_style,
    downcase_first_character,
    set_text_style,
    set_alignment,
)
from .models import MenuDay

FIRST_NORMS_ANALYSIS_ROW = 8

logger = logging.getLogger(__name__)


@task
def create_norms_analysis_report(start_date, end_date, id_group):
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    menu_days = MenuDay.objects.filter(
        Q(date__gte=start_date),
        Q(date__lte=end_date),
        Q(status=MenuDay.STATUS.published),
    ).order_by('date')
    menu_days = [obj for obj in menu_days if obj.group.id == id_group]

    product_categories = {}  # type: Dict[Any, Any]
    for menu_day in menu_days:
        persons_numbers = list(
            menu_day.numbers_of_persons.order_by('age_category__sort').values_list(
                'value', flat=True
            )
        )
        menu_day_key = menu_day.date.strftime('%d.%m.%Y')

        dish_items = menu_day.dish_items.order_by('period')
        for dish_item in dish_items:
            # dishes
            for dish in dish_item.dishes.iterator():
                # dish products
                for product in dish.products.iterator():
                    category_name = product.category.name
                    product_name = product.name

                    if category_name not in product_categories:
                        product_categories[category_name] = {}
                        product_categories[category_name]['norms'] = list(
                            product.category.product_norms.filter(group__pk=id_group)
                            .order_by('age_category__sort')
                            .values_list('value', flat=True)
                        )

                    if product_name not in product_categories[category_name]:
                        # https://stackoverflow.com/questions/15516413/dict-fromkeys-all-point-to-same-list
                        product_categories[category_name][product_name] = dict.fromkeys(
                            [obj.date.strftime('%d.%m.%Y') for obj in menu_days]
                        )
                        for k, _ in product_categories[category_name][
                            product_name
                        ].items():
                            product_categories[category_name][product_name][k] = []

                    # product norms
                    product_norms = dish.norms.filter(
                        group__pk=id_group, product=product
                    ).order_by('age_category__sort')
                    index = 0
                    calculated_products_norms = []
                    for norm in product_norms.iterator():
                        calculated_products_norms.append(
                            norm.value
                            * (1 if persons_numbers[index] > 0 else 0)
                            / product.convert_ratio  # per one person
                        )
                        index += 1
                    product_day_dict = product_categories[category_name][product_name][
                        menu_day_key
                    ]
                    if len(product_day_dict) > 0:
                        product_categories[category_name][product_name][
                            menu_day_key
                        ] = [
                            x + y
                            for x, y in zip(product_day_dict, calculated_products_norms)
                        ]
                    else:
                        product_categories[category_name][product_name][
                            menu_day_key
                        ] = calculated_products_norms

    report_path = 'spreadsheets/norms_analysis_{}d.xlsx'
    menu_days_count = len(menu_days)
    report_path = (
        report_path.format(20) if menu_days_count <= 20 else report_path.format(31)
    )
    abs_report_path = str(settings.ROOT_DIR.path(report_path))
    workbook = load_workbook(abs_report_path)

    group = Group.objects.get(pk=id_group)
    age_categories = list(group.age_categories.values_list('name', flat=True))

    # create worksheets
    iter_age_categories = iter(age_categories)
    worksheet = workbook.active
    worksheet.title = next(iter_age_categories)
    for age_category in iter_age_categories:
        target = workbook.copy_worksheet(worksheet)
        target.title = age_category

    for worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]

        translation.activate('uk')
        worksheet['A3'].value = worksheet['A3'].value.format(
            month=date_format(start_date, 'F').upper(),
            year=end_date.strftime('%Y'),
            group_name=group.name,
            age_category_name=worksheet_name,
        )

    # sort by product and categoryy
    for category_name, category in product_categories.items():
        sorted_products = dict(sorted(category.items()))
        product_categories[category_name] = sorted_products
    sorted_product_categories = dict(sorted(product_categories.items()))

    numbers_of_rows = fill_norms_analysis_workbook(workbook, sorted_product_categories)
    format_norms_analysis_workbook(
        workbook, numbers_of_rows, menu_days_count, sorted_product_categories
    )

    group_name = group.name.replace(', ', '_').replace(' ', '_')
    report_file = '{month}_{group}.xlsx'.format(
        month=start_date.strftime('%Y-%m'), group=group_name
    )
    workbook.save(f"{settings.REPORTS_DIR}/{report_file}")
    results = {'report_path': f"{settings.MEDIA_URL}reports/{report_file}"}

    report_file = 'test_analysis.json'
    report_path = f"{settings.REPORTS_DIR}/{report_file}"
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(
            sorted_product_categories, f, ensure_ascii=False, default=str, indent=4
        )
    logger.info('Report of norms analysis is created.')

    return results


def fill_norms_analysis_workbook(workbook, product_categories):
    worksheet_index = 0
    for worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]

        rownum = 1
        index = 0
        for category_name, category in product_categories.items():
            worksheet.cell(row=FIRST_NORMS_ANALYSIS_ROW + index, column=1, value=rownum)
            worksheet.cell(
                row=FIRST_NORMS_ANALYSIS_ROW + index, column=2, value=category_name
            )
            worksheet.cell(
                row=FIRST_NORMS_ANALYSIS_ROW + index,
                column=3,
                value=category['norms'][worksheet_index] * 1000,
            )
            category_index = index

            rownum += 1
            index += 1
            for product_name, product in category.items():
                if product_name == 'norms':
                    continue
                worksheet.cell(
                    row=FIRST_NORMS_ANALYSIS_ROW + index,
                    column=2,
                    value=downcase_first_character(product_name),
                )
                col_index = 4
                day_count = 0
                for day, issued_values in product.items():
                    if day_count == 10:
                        col_index += 2
                    elif day_count == 20:
                        col_index += 2
                    worksheet.cell(
                        row=FIRST_NORMS_ANALYSIS_ROW - 2, column=col_index, value=day
                    )
                    worksheet.cell(
                        row=FIRST_NORMS_ANALYSIS_ROW + index,
                        column=col_index,
                        value=issued_values[worksheet_index] * 1000
                        if len(issued_values) > 0
                        else 0,
                    )
                    col_index += 1
                    day_count += 1
                index += 1

            for j in range(day_count):
                if j < 10:
                    col_index = j
                elif 10 <= j < 20:
                    col_index = j + 2
                elif j >= 20:
                    col_index = j + 4
                total_issued_by_day = worksheet.cell(
                    row=FIRST_NORMS_ANALYSIS_ROW + category_index, column=4 + col_index
                )
                total_issued_by_day.value = '=SUM({}{}:{}{})'.format(
                    total_issued_by_day.column_letter,
                    FIRST_NORMS_ANALYSIS_ROW + 1 + category_index,
                    total_issued_by_day.column_letter,
                    FIRST_NORMS_ANALYSIS_ROW - 1 + index,
                )
            category_row = FIRST_NORMS_ANALYSIS_ROW + category_index
            average_by_days = worksheet.cell(row=category_row, column=4 + 10)
            average_by_days.value = '=IFERROR(AVERAGEIF(D{}:M{},">=0"),0)'.format(
                category_row, category_row
            )
            average_by_days = worksheet.cell(row=category_row, column=4 + 22)
            average_by_days.value = '=IFERROR(AVERAGEIF(P{}:Y{},">=0"),0)'.format(
                category_row, category_row
            )
            delta_result = worksheet.cell(row=category_row, column=4 + 11)
            delta_result.value = '=N{}-C{}'.format(category_row, category_row)
            delta_result = worksheet.cell(row=category_row, column=4 + 23)
            delta_result.value = '=Z{}-C{}'.format(category_row, category_row)

            if day_count > 20:
                average_by_days = worksheet.cell(row=category_row, column=4 + 35)
                average_by_days.value = '=IFERROR(AVERAGEIF(AB{}:AL{},">=0"),0)'.format(
                    category_row, category_row
                )
                delta_result = worksheet.cell(row=category_row, column=4 + 36)
                delta_result.value = '=AM{}-C{}'.format(category_row, category_row)

                average_month_result = worksheet.cell(row=category_row, column=4 + 37)
                average_month_result.value = '=IFERROR(AVERAGE(N{},Z{},AM{}),0)'.format(
                    category_row, category_row, category_row
                )
                delta_month_result = worksheet.cell(row=category_row, column=4 + 38)
                delta_month_result.value = '=AO{}-C{}'.format(
                    category_row, category_row
                )
                completion_month = worksheet.cell(row=category_row, column=4 + 39)
                completion_month.value = '=IFERROR(AO{}*100/C{},0)'.format(
                    category_row, category_row
                )
            else:

                average_month_result = worksheet.cell(row=category_row, column=4 + 24)
                average_month_result.value = '=IFERROR(AVERAGE(N{},Z{}),0)'.format(
                    category_row, category_row
                )
                delta_month_result = worksheet.cell(row=category_row, column=4 + 25)
                delta_month_result.value = '=AB{}-C{}'.format(
                    category_row, category_row
                )
                completion_month = worksheet.cell(row=category_row, column=4 + 26)
                completion_month.value = '=IFERROR(AB{}*100/C{},0)'.format(
                    category_row, category_row
                )

        worksheet_index += 1

    return index


def format_norms_analysis_workbook(
    workbook, number_of_rows, menu_days_count, product_categories
):
    worksheet = workbook.active
    base_page_setup = copy(worksheet.page_setup)
    base_page_margins = copy(worksheet.page_margins)
    base_sheet_properties = copy(worksheet.sheet_properties)
    base_print_title_cols = copy(worksheet.print_title_cols)
    base_print_title_rows = copy(worksheet.print_title_rows)

    for worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]

        first_cell = worksheet.cell(row=FIRST_NORMS_ANALYSIS_ROW, column=1)
        number_of_columns = 30 if menu_days_count <= 20 else 43
        for i in range(number_of_rows):
            for j in range(number_of_columns):
                active_cell = worksheet.cell(
                    row=first_cell.row + i, column=first_cell.column + j
                )
                clone_cell_style(first_cell, active_cell)
                if j in (2, 13, 14, 25, 26, 38, 39) or j > (number_of_columns - 4):
                    active_cell.alignment = Alignment(
                        horizontal='center', vertical='top'
                    )
                else:
                    active_cell.alignment = Alignment(vertical='top')
                if j in (13, 14, 25, 26, 38, 39) or j > (number_of_columns - 4):
                    active_cell.number_format = '0.0'

        index = 0
        for _, category in product_categories.items():
            cell = worksheet.cell(row=FIRST_NORMS_ANALYSIS_ROW + index, column=2)
            set_text_style(cell, bold=True)
            index += 1

            for product_name, product in category.items():
                if product_name == 'norms':
                    continue
                cell = worksheet.cell(row=FIRST_NORMS_ANALYSIS_ROW + index, column=2)
                set_text_style(cell, italic=True)
                set_alignment(cell, 2)
                index += 1

        # clone worksheet settings
        worksheet.page_setup = base_page_setup
        worksheet.page_margins = base_page_margins
        worksheet.sheet_properties = base_sheet_properties
        worksheet.print_title_cols = base_print_title_cols
        worksheet.print_title_rows = base_print_title_rows
